import os
import pickle

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

GATE_FILLS = {
    'H':    'D6E4F0',
    'I':    'E8E8E8',
    'CNOT': 'F5C4A1',
    'RX':   'C6E8D4',
    'RY':   'C6E8D4',
    'RZ':   'C6E8D4',
}


def _sanitize_sheet_name(name: str) -> str:
    for ch in ['\\', '/', '*', '[', ']', ':', '?']:
        name = name.replace(ch, '_')
    return name[:31]


def list_experiments_with_candidates(results_dir: str) -> list[str]:
    """Finds every subfolder of results_dir that has a saved candidate.pkl
    (this naturally excludes Classical_baseline, which has no candidate)."""
    if not os.path.isdir(results_dir):
        return []
    return sorted(
        entry for entry in os.listdir(results_dir)
        if os.path.isfile(os.path.join(results_dir, entry, 'candidate.pkl'))
    )


def gene_table(name: str, results_dir: str) -> list[dict]:
    """Builds (gene, layer, qubit, gate, angle) rows for one experiment."""
    path = os.path.join(results_dir, name, 'candidate.pkl')
    with open(path, 'rb') as f:
        data = pickle.load(f)
    candidate, genes = data['candidate'], data['genes']

    rows = []
    for i, info in enumerate(genes):
        layer, qubit = divmod(i, candidate.n_qubits)
        angle = info.get('angle', '')
        rows.append({
            'gene': f'g{i + 1}', 'layer': layer, 'qubit': qubit,
            'gate': info['gate'],
            'angle': round(angle, 4) if isinstance(angle, (int, float)) else angle,
        })
    return rows


def export_gene_layouts_to_excel(results_dir: str, output_path: str):
    """
    Scans results_dir for every experiment with a saved candidate and writes
    one Excel sheet per experiment, each with its full gene layout table.
    """
    names = list_experiments_with_candidates(results_dir)
    if not names:
        raise ValueError(f"No saved candidates found under '{results_dir}'")

    wb = Workbook()
    wb.remove(wb.active)

    headers = ['gene', 'layer', 'qubit', 'gate', 'angle']
    bold, center = Font(bold=True), Alignment(horizontal='center')

    for name in names:
        ws = wb.create_sheet(_sanitize_sheet_name(name))

        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = bold
            cell.alignment = center

        for r, row in enumerate(gene_table(name, results_dir), start=2):
            for c, h in enumerate(headers, start=1):
                cell = ws.cell(row=r, column=c, value=row[h])
                cell.alignment = center
                if h == 'gate' and row[h] in GATE_FILLS:
                    cell.fill = PatternFill('solid', start_color=GATE_FILLS[row[h]])

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12
        ws.freeze_panes = 'A2'

    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    wb.save(output_path)
    print(f"Saved → {output_path}  ({len(names)} sheets)")